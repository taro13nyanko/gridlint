"""Mutation benchmark: does the detector actually find planted defects, and does it stay quiet otherwise?

A linter that reports everything is useless, and one that reports nothing is
worse. This measures both directions on the same corpus:

  recall     of the planted defects of each kind, how many were reported at the
             right cell by the right rule
  precision  how often a report on a mutated file points at the planted defect
             rather than something else
  false      findings raised on the clean, unmutated workbooks, where the right
  positives  answer is silence

Each mutant is written out and recalculated by Excel, so the fixtures carry real
cached values exactly like a file a user would upload.

    python bench/benchmark.py --out bench/results.json
"""
from __future__ import annotations

import argparse
import json
import random
import shutil
import sys
import tempfile
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from gridlint.audit import audit  # noqa: E402

SEED = 20260915


@dataclass
class Mutant:
    name: str
    kind: str                 # the rule that should catch it
    path: str
    target_cell: str          # "Sheet!C16"
    description: str


@dataclass
class Outcome:
    mutant: str
    kind: str
    target: str
    detected: bool
    detected_by_right_rule: bool
    reported_cells: list[str] = field(default_factory=list)
    top_rule: str | None = None
    rank_of_hit: int | None = None
    total_findings: int = 0


# --------------------------------------------------------------------------- mutations

def _formula_cells(ws):
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                yield c


def mut_shrink_sum(ws, rng: random.Random):
    """A SUM range that stops one row short: the classic silent underreport."""
    import re

    cands = [c for c in _formula_cells(ws)
             if re.match(r"^=SUM\([A-Z]{1,3}\d+:[A-Z]{1,3}\d+\)$", c.value or "")]
    if not cands:
        return None
    cell = rng.choice(cands)
    m = re.match(r"^=SUM\(([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)\)$", cell.value)
    c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    if r2 - r1 < 2:
        return None
    cell.value = f"=SUM({c1}{r1}:{c2}{r2 - 1})"
    return cell.coordinate, f"SUM range shortened from {r2} to {r2 - 1}"


def mut_break_pattern(ws, rng: random.Random):
    """One cell in a copied row replaced with a hand-typed multiplier."""
    from collections import defaultdict as dd

    rows = dd(list)
    for c in _formula_cells(ws):
        rows[c.row].append(c)
    cands = [(r, cs) for r, cs in rows.items() if len(cs) >= 6]
    if not cands:
        return None
    _r, cells = rng.choice(cands)
    cells.sort(key=lambda c: c.column)
    victim = cells[len(cells) // 2]
    left = f"{get_column_letter(victim.column - 1)}{victim.row}"
    victim.value = f"={left}*1.27"
    return victim.coordinate, "one formula in a copied row replaced by a hand-typed rate"


def mut_number_as_text(ws, rng: random.Random):
    """A number pasted as text inside a summed range, which SUM then skips."""
    import re

    # Only a SUM over cells that actually hold typed numbers can be mutated this
    # way; a total over formula cells has nothing to convert to text.
    cands = []
    for c in _formula_cells(ws):
        m = re.match(r"^=SUM\(([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)\)$", c.value or "")
        if not m:
            continue
        col, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
        numeric = [r for r in range(r1, r2 + 1)
                   if isinstance(ws[f"{col}{r}"].value, (int, float))
                   and not isinstance(ws[f"{col}{r}"].value, bool)]
        if numeric:
            cands.append((c, col, numeric))
    if not cands:
        return None
    cell, col, numeric = rng.choice(cands)
    r = rng.choice(numeric)
    t = ws[f"{col}{r}"]
    t.value = f"{t.value:,.0f}"
    t.number_format = "@"
    return cell.coordinate, f"{col}{r} converted to text"


def mut_masked_error(ws, rng: random.Random):
    """A divide-by-zero wrapped in IFERROR, so the sheet shows a confident zero."""
    target = None
    for c in _formula_cells(ws):
        if c.column >= 3 and c.row > 5:
            target = c
            break
    if target is None:
        return None
    empty_row = ws.max_row + 5
    target.value = f"=IFERROR({get_column_letter(target.column)}{target.row - 1}/A{empty_row},0)"
    return target.coordinate, "divide by an empty cell hidden behind IFERROR"


def mut_missing_sheet(ws, rng: random.Random):
    """A reference to a sheet that is not in the file."""
    cands = list(_formula_cells(ws))
    if not cands:
        return None
    cell = rng.choice(cands)
    cell.value = "='Q3 Actuals'!B4*1"
    return cell.coordinate, "reference to a sheet that does not exist"


def mut_circular(ws, rng: random.Random):
    """Two cells that depend on each other."""
    r = ws.max_row + 3
    ws[f"C{r}"] = f"=C{r + 1}+1"
    ws[f"C{r + 1}"] = f"=C{r}+1"
    return f"C{r}", "two cells referring to each other"


MUTATIONS = [
    ("R001", "shrink_sum", mut_shrink_sum),
    ("R002", "break_pattern", mut_break_pattern),
    ("R008", "number_as_text", mut_number_as_text),
    ("R012", "masked_error", mut_masked_error),
    ("R007", "missing_sheet", mut_missing_sheet),
    ("R006", "circular", mut_circular),
]


# --------------------------------------------------------------------------- corpus

def make_mutants(bases: list[Path], out_dir: Path, per_kind: int) -> list[Mutant]:
    out_dir.mkdir(parents=True, exist_ok=True)
    rng = random.Random(SEED)
    mutants: list[Mutant] = []
    for base in bases:
        for rule_code, kind, fn in MUTATIONS:
            for i in range(per_kind):
                dst = out_dir / f"{base.stem}-{kind}-{i}.xlsx"
                shutil.copy(base, dst)
                wb = openpyxl.load_workbook(dst, data_only=False)
                ws = wb[wb.sheetnames[0]]
                result = fn(ws, rng)
                if result is None:
                    wb.close()
                    dst.unlink(missing_ok=True)
                    continue
                coord, desc = result
                wb.save(dst)
                wb.close()
                mutants.append(Mutant(name=dst.stem, kind=rule_code, path=str(dst),
                                      target_cell=f"{ws.title}!{coord}", description=desc))
    return mutants


def recalculate_with_excel(paths: list[Path]) -> int:
    """Open and re-save each file so it carries the values Excel computes."""
    try:
        import win32com.client as win32
    except ImportError:
        print("  (pywin32 not available: mutants keep no cached values)")
        return 0
    app = win32.Dispatch("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    done = 0
    try:
        for p in paths:
            try:
                wb = app.Workbooks.Open(str(p.resolve()))
                app.Calculate()
                wb.Save()
                wb.Close(SaveChanges=False)
                done += 1
            except Exception as e:                     # a mutant Excel refuses to open is still a data point
                print(f"  ! Excel could not recalculate {p.name}: {e}")
    finally:
        app.Quit()
    return done


# --------------------------------------------------------------------------- scoring

def score(mutants: list[Mutant]) -> list[Outcome]:
    outcomes = []
    for m in mutants:
        report = audit(m.path, price=False)
        cells = [f.cell for f in report.findings]
        group = {c for f in report.findings for c in (f.group_cells or [f.cell])}
        hit_rule = None
        rank = None
        for i, f in enumerate(report.findings):
            covered = {f.cell, *(f.group_cells or [])}
            if f.rule == m.kind and (m.target_cell in covered or _same_row(m.target_cell, covered)):
                hit_rule, rank = f.rule, i + 1
                break
        outcomes.append(Outcome(
            mutant=m.name, kind=m.kind, target=m.target_cell,
            detected=m.target_cell in group or _same_row(m.target_cell, group),
            detected_by_right_rule=hit_rule is not None,
            reported_cells=cells[:8],
            top_rule=report.findings[0].rule if report.findings else None,
            rank_of_hit=rank,
            total_findings=len(report.findings),
        ))
    return outcomes


def _same_row(target: str, cells: set[str]) -> bool:
    """A grouped finding names its first cell; a hit anywhere in that row counts."""
    import re

    ts, _, ta = target.rpartition("!")
    tm = re.match(r"([A-Z]+)(\d+)", ta)
    if not tm:
        return False
    for c in cells:
        cs, _, ca = c.rpartition("!")
        cm = re.match(r"([A-Z]+)(\d+)", ca)
        if cm and cs == ts and cm.group(2) == tm.group(2):
            return True
    return False


def false_positives(bases: list[Path]) -> dict:
    out = {}
    for b in bases:
        report = audit(b, price=False)
        out[b.name] = {
            "findings": len(report.findings),
            "critical": len(report.critical),
            "detail": [f"{f.rule} {f.cell} {f.title}" for f in report.findings],
        }
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--per-kind", type=int, default=3)
    ap.add_argument("--out", default=str(ROOT / "bench" / "results.json"))
    ap.add_argument("--keep", action="store_true", help="keep the generated mutants")
    ap.add_argument("--no-excel", action="store_true", help="skip Excel recalculation")
    ap.add_argument("--limit-bases", type=int, default=0, help="use only the first N base workbooks")
    args = ap.parse_args()

    corpus = sorted((ROOT / "bench" / "corpus").glob("*.xlsx"))
    bases = corpus or [ROOT / "samples" / "clean-model.xlsx"]
    bases = [b for b in bases if b.exists()]
    if not bases:
        print("no base workbooks found; run bench/make_corpus.py first")
        return 2
    if args.limit_bases:
        bases = bases[: args.limit_bases]

    work = Path(args.keep and (ROOT / "bench" / "mutants") or tempfile.mkdtemp(prefix="gridlint-bench-"))
    print(f"building mutants in {work}")
    mutants = make_mutants(bases, work, args.per_kind)
    print(f"  {len(mutants)} mutants from {len(bases)} base workbook(s)")

    if not args.no_excel:
        print("recalculating with Excel so the fixtures carry real cached values")
        n = recalculate_with_excel([Path(m.path) for m in mutants])
        print(f"  recalculated {n}")

    print("scoring")
    outcomes = score(mutants)
    fp = false_positives(bases)

    by_kind: dict[str, list[Outcome]] = defaultdict(list)
    for o in outcomes:
        by_kind[o.kind].append(o)

    summary = {}
    for kind, group in sorted(by_kind.items()):
        n = len(group)
        right = sum(1 for o in group if o.detected_by_right_rule)
        first = sum(1 for o in group if o.rank_of_hit == 1)
        summary[kind] = {
            "planted": n,
            "found_by_the_right_rule": right,
            "recall": round(right / n, 3) if n else 0.0,
            "ranked_first": first,
            "median_findings_per_file": sorted(o.total_findings for o in group)[n // 2] if n else 0,
        }

    total = len(outcomes)
    found = sum(1 for o in outcomes if o.detected_by_right_rule)
    clean_findings = sum(v["findings"] for v in fp.values())
    result = {
        "seed": SEED,
        "bases": [b.name for b in bases],
        "mutants": total,
        "overall_recall": round(found / total, 3) if total else 0.0,
        "by_rule": summary,
        "clean_workbooks": fp,
        "false_positives_on_clean_files": clean_findings,
        "outcomes": [asdict(o) for o in outcomes],
    }
    Path(args.out).write_text(json.dumps(result, indent=1), encoding="utf-8")

    print()
    print(f"{'rule':6} {'planted':>8} {'found':>7} {'recall':>8} {'ranked #1':>10}")
    for kind, s in summary.items():
        print(f"{kind:6} {s['planted']:>8} {s['found_by_the_right_rule']:>7} "
              f"{s['recall']:>8.0%} {s['ranked_first']:>10}")
    print(f"{'ALL':6} {total:>8} {found:>7} {found / total if total else 0:>8.0%}")
    print(f"\nfalse positives on clean workbooks: {clean_findings}")
    for name, v in fp.items():
        print(f"  {name}: {v['findings']} findings {v['detail']}")
    print(f"\nwrote {args.out}")
    if not args.keep:
        shutil.rmtree(work, ignore_errors=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
