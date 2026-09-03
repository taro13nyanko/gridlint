"""How large a workbook Gridlint can take, and where the time goes.

    python bench/perf.py --rows 2000 --cols 12

Builds a synthetic model of a given size, then times each stage separately so
the answer is useful rather than a single number: parsing, graph construction,
recalculation, rule evaluation, and pricing the fixes.

The generated file has no cached values, so this measures the engine, not Excel.
"""
from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from gridlint import engine, workbook  # noqa: E402
from gridlint.audit import audit  # noqa: E402
from gridlint.rules import registry  # noqa: E402


def build(path: Path, rows: int, cols: int) -> int:
    """A wide model: line items growing across columns, subtotals, a grand total."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Line"
    ws["B1"] = "Growth"
    for c in range(cols):
        ws.cell(row=1, column=3 + c, value=f"M{c + 1}")

    formulas = 0
    section_totals: list[int] = []
    r = 2
    section_start = r
    for i in range(rows):
        ws.cell(row=r, column=1, value=f"Line item {i + 1}")
        ws.cell(row=r, column=2, value=0.01 + (i % 7) / 100)
        ws.cell(row=r, column=3, value=1000 + i)
        for c in range(1, cols):
            col = get_column_letter(3 + c)
            prev = get_column_letter(2 + c)
            ws.cell(row=r, column=3 + c).value = f"={prev}{r}*(1+$B${r})"
            formulas += 1
        r += 1
        if (i + 1) % 25 == 0:                     # a subtotal every 25 line items
            ws.cell(row=r, column=1, value="Subtotal")
            for c in range(cols):
                col = get_column_letter(3 + c)
                ws.cell(row=r, column=3 + c).value = f"=SUM({col}{section_start}:{col}{r - 1})"
                formulas += 1
            section_totals.append(r)
            r += 2
            section_start = r

    ws.cell(row=r, column=1, value="Grand total")
    for c in range(cols):
        col = get_column_letter(3 + c)
        refs = ",".join(f"{col}{t}" for t in section_totals) or f"{col}2"
        ws.cell(row=r, column=3 + c).value = f"=SUM({refs})"
        formulas += 1
    grand = r
    r += 2
    ws.cell(row=r, column=1, value="Full year")
    ws.cell(row=r, column=3).value = f"=SUM(C{grand}:{get_column_letter(2 + cols)}{grand})"
    formulas += 1

    wb.save(path)
    wb.close()
    return formulas


def timed(label: str, fn):
    t0 = time.perf_counter()
    out = fn()
    return out, (time.perf_counter() - t0) * 1000


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", type=int, default=2000)
    ap.add_argument("--cols", type=int, default=12)
    ap.add_argument("--out", default=str(ROOT / "bench" / "perf.json"))
    args = ap.parse_args()

    path = ROOT / "bench" / f"perf-{args.rows}x{args.cols}.xlsx"
    n = build(path, args.rows, args.cols)
    size_kb = path.stat().st_size / 1024
    print(f"built {path.name}: {n:,} formulas, {size_kb:,.0f} KB")

    # One throwaway pass so the numbers are steady-state rather than a
    # measurement of Python's import and regex-compilation cost.
    warm = workbook.load(path)
    engine.recalc(warm, engine.build_graph(warm))

    wb, t_load = timed("load", lambda: workbook.load(path))
    graph, t_graph = timed("graph", lambda: engine.build_graph(wb))
    computed, t_calc = timed("recalc", lambda: engine.recalc(wb, graph))

    def run_rules():
        found = []
        for meta, fn in registry():
            found.extend(fn(wb=wb, computed=computed, graph=graph, self_check=None))
        return found

    findings, t_rules = timed("rules", run_rules)
    report, t_total = timed("audit", lambda: audit(path))

    stats = wb.stats()
    result = {
        "rows": args.rows, "cols": args.cols,
        "formulas": stats["formulas"], "cells": stats["cells"], "file_kb": round(size_kb),
        "ms": {
            "load_and_parse": round(t_load),
            "build_graph": round(t_graph),
            "recalculate": round(t_calc),
            "run_rules": round(t_rules),
            "full_audit_including_pricing": round(t_total),
        },
        "formulas_per_second": round(stats["formulas"] / max(t_calc / 1000, 1e-6)),
        "findings": len(report.findings),
        "graph_edges": sum(len(v) for v in graph.dependents.values()),
    }
    Path(args.out).write_text(json.dumps(result, indent=1), encoding="utf-8")

    print()
    print(f"{'stage':34} {'ms':>8}")
    for k, v in result["ms"].items():
        print(f"{k:34} {v:>8,}")
    print()
    print(f"{stats['formulas']:,} formulas, {result['graph_edges']:,} graph edges")
    print(f"recalculation: {result['formulas_per_second']:,} formulas/second")
    print(f"findings: {result['findings']}")
    print()
    print("Steady state, after one warm-up pass. A cold first run in a fresh")
    print("process adds a few seconds of imports and regex compilation.")
    print(f"wrote {args.out}")
    path.unlink(missing_ok=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
