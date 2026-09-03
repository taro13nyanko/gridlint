"""Write a corrected copy of a workbook.

Gridlint never edits the file it was given. It writes a new one, keeps the
original formulas in a cell comment, and leaves the corrected cells highlighted
so a reviewer can see exactly what moved.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

from .rules.base import Finding

CHANGED_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")


def apply_fixes(src: str | Path, findings: Iterable[Finding], dest: str | Path) -> int:
    """Apply the verified fixes to a copy of `src`. Returns the number of cells changed."""
    wb = openpyxl.load_workbook(src, data_only=False, keep_vba=str(src).endswith(".xlsm"))
    changed = 0
    for f in findings:
        if not f.fix:
            continue
        for edit in f.fix.edits:
            sheet_name, _, a1 = edit.cell.rpartition("!")
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if ws is None:
                continue
            cell = ws[a1]
            old = cell.value
            cell.value = edit.new_formula
            cell.fill = CHANGED_FILL
            note = (f"Gridlint {f.rule}: {f.title}\n"
                    f"was: {old}\n"
                    f"now: {edit.new_formula}\n"
                    f"{f.detail}")
            cell.comment = Comment(note[:2000], "Gridlint")
            changed += 1
    Path(dest).parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    wb.close()
    return changed
