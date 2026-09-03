"""Load an .xlsx into a model Gridlint can analyse and recalculate.

Every workbook saved by Excel, LibreOffice or Google Sheets carries the values
those apps last computed. Gridlint loads the file twice -- once for formulas,
once for those cached values -- which gives an independent ground truth to
check its own calculation engine against before it dares suggest an edit.
"""
from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from .formula.ast_nodes import ExternalRef, Node, TableRef, walk
from .formula.parser import parse_formula
from .formula.tokenizer import FormulaSyntaxError
from .formula.values import BLANK, ExcelError

ERROR_LITERALS = {"#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#NULL!", "#N/A", "#SPILL!", "#CALC!"}


def a1(col: int, row: int) -> str:
    return f"{get_column_letter(col)}{row}"


@dataclass
class TableInfo:
    """An Excel table, so Table1[Amount] can be resolved to a real range."""
    name: str
    sheet: str
    col1: int
    row1: int
    col2: int
    row2: int
    columns: list[str] = field(default_factory=list)
    header_rows: int = 1
    totals_rows: int = 0

    def column_range(self, column: str | None) -> tuple[int, int, int, int]:
        """(col1, row1, col2, row2) for one column's data rows, or the whole body."""
        first_data = self.row1 + self.header_rows
        last_data = self.row2 - self.totals_rows
        if column is None:
            return self.col1, first_data, self.col2, last_data
        wanted = column.strip().lower()
        for i, name in enumerate(self.columns):
            if name.strip().lower() == wanted:
                c = self.col1 + i
                return c, first_data, c, last_data
        raise KeyError(column)


@dataclass
class Cell:
    sheet: str
    col: int
    row: int
    formula: str | None = None
    cached: Any = None              # what Excel last computed (None if never computed)
    static: Any = None              # literal value for non-formula cells
    ast: Node | None = None
    parse_error: str | None = None
    number_format: str = "General"
    is_array: bool = False
    uses_external: bool = False       # reads a workbook this file does not contain
    uses_table: bool = False

    @property
    def addr(self) -> str:
        return a1(self.col, self.row)

    @property
    def ref(self) -> str:
        return f"{self.sheet}!{self.addr}"

    def __repr__(self) -> str:  # pragma: no cover - debugging aid
        return f"<Cell {self.ref} {self.formula or self.static!r}>"


@dataclass
class Sheet:
    name: str
    index: int
    cells: dict[tuple[int, int], Cell] = field(default_factory=dict)
    max_col: int = 0
    max_row: int = 0
    hidden: bool = False
    hidden_rows: set[int] = field(default_factory=set)
    hidden_cols: set[int] = field(default_factory=set)

    def cell(self, col: int, row: int) -> Cell | None:
        return self.cells.get((col, row))


@dataclass
class Workbook:
    path: str
    sheets: dict[str, Sheet] = field(default_factory=dict)
    defined_names: dict[str, str] = field(default_factory=dict)
    tables: dict[str, TableInfo] = field(default_factory=dict)
    load_warnings: list[str] = field(default_factory=list)

    @property
    def sheet_list(self) -> list[Sheet]:
        return sorted(self.sheets.values(), key=lambda s: s.index)

    def formula_cells(self) -> Iterator[Cell]:
        for s in self.sheet_list:
            for (_c, _r), cell in sorted(s.cells.items(), key=lambda kv: (kv[0][1], kv[0][0])):
                if cell.formula:
                    yield cell

    def all_cells(self) -> Iterator[Cell]:
        for s in self.sheet_list:
            for cell in s.cells.values():
                yield cell

    def get(self, sheet: str, col: int, row: int) -> Cell | None:
        s = self.sheets.get(sheet)
        return s.cell(col, row) if s else None

    @property
    def has_cached_values(self) -> bool:
        return any(c.cached is not None for c in self.formula_cells())

    def stats(self) -> dict[str, int]:
        formulas = list(self.formula_cells())
        return {
            "sheets": len(self.sheets),
            "cells": sum(len(s.cells) for s in self.sheets.values()),
            "formulas": len(formulas),
            "unparsed": sum(1 for c in formulas if c.parse_error),
            "tables": len(self.tables),
            "external_links": sum(1 for c in formulas if c.uses_external),
        }


def _norm(v: Any) -> Any:
    """Normalise an openpyxl value into the evaluator's value model."""
    if v is None:
        return None
    if isinstance(v, str):
        if v in ERROR_LITERALS:
            return ExcelError(v)
        return v
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time, _dt.timedelta)):
        return v
    return str(v)


def load(path: str | Path) -> Workbook:
    p = Path(path)
    wb_f = openpyxl.load_workbook(p, data_only=False, read_only=False, keep_links=False)
    wb_v = openpyxl.load_workbook(p, data_only=True, read_only=False, keep_links=False)
    out = Workbook(path=str(p))

    for name, dn in (getattr(wb_f, "defined_names", {}) or {}).items():
        try:
            out.defined_names[str(name).upper()] = str(dn.value)
        except Exception:  # pragma: no cover - malformed names are not fatal
            continue

    for idx, ws in enumerate(wb_f.worksheets):
        sheet = Sheet(name=ws.title, index=idx, hidden=ws.sheet_state != "visible")
        # openpyxl's TableList.items() yields (name, ref string); the Table
        # objects with their columns are only reachable through .values().
        for table in _worksheet_tables(ws):
            info = _table_info(ws.title, table)
            if info is not None:
                out.tables[info.name.lower()] = info
        for r_index, dim in (ws.row_dimensions or {}).items():
            if getattr(dim, "hidden", False):
                sheet.hidden_rows.add(int(r_index))
        from openpyxl.utils import column_index_from_string
        for c_key, dim in (ws.column_dimensions or {}).items():
            if not getattr(dim, "hidden", False):
                continue
            try:
                start = column_index_from_string(getattr(dim, "min", None) and
                                                 get_column_letter(dim.min) or str(c_key))
                end = int(getattr(dim, "max", start) or start)
                sheet.hidden_cols.update(range(int(start), int(end) + 1))
            except Exception:
                continue
        vs = wb_v[ws.title]
        for row in ws.iter_rows():
            for c in row:
                raw = c.value
                if raw is None:
                    continue
                cached = _norm(vs.cell(row=c.row, column=c.column).value)
                is_array = isinstance(raw, ArrayFormula)
                if is_array:
                    raw = raw.text
                if isinstance(raw, str) and raw.startswith("="):
                    cell = Cell(sheet=ws.title, col=c.column, row=c.row, formula=raw,
                                cached=cached, number_format=c.number_format or "General",
                                is_array=is_array)
                    try:
                        cell.ast = parse_formula(raw)
                        cell.uses_external = any(isinstance(n, ExternalRef) for n in walk(cell.ast))
                        cell.uses_table = any(isinstance(n, TableRef) for n in walk(cell.ast))
                    except FormulaSyntaxError as e:
                        cell.parse_error = str(e)
                    except RecursionError:
                        cell.parse_error = "formula too deeply nested"
                else:
                    cell = Cell(sheet=ws.title, col=c.column, row=c.row, static=_norm(raw),
                                cached=cached, number_format=c.number_format or "General")
                sheet.cells[(c.column, c.row)] = cell
                sheet.max_col = max(sheet.max_col, c.column)
                sheet.max_row = max(sheet.max_row, c.row)
        out.sheets[ws.title] = sheet

    wb_f.close()
    wb_v.close()
    if not out.sheets:
        out.load_warnings.append("workbook contains no worksheets")
    return out


def _worksheet_tables(ws: Any) -> list[Any]:
    tables = getattr(ws, "tables", None)
    if not tables:
        return []
    try:
        return [t for t in tables.values() if hasattr(t, "ref")]
    except Exception:                      # pragma: no cover - defensive
        return []


def _table_info(sheet: str, table: Any) -> TableInfo | None:
    """Read one openpyxl table definition into something the engine can resolve."""
    ref = getattr(table, "ref", None)
    if not ref or ":" not in ref:
        return None
    try:
        from .formula.parser import parse_range
        rng = parse_range(ref)
    except Exception:
        return None
    columns = [str(getattr(c, "name", "")) for c in (getattr(table, "tableColumns", None) or [])]
    display = str(getattr(table, "displayName", "") or getattr(table, "name", "") or "")
    if not display:
        return None
    return TableInfo(
        name=display, sheet=sheet,
        col1=rng.col1, row1=rng.row1, col2=rng.col2, row2=rng.row2,
        columns=columns,
        header_rows=int(getattr(table, "headerRowCount", 1) or 0),
        totals_rows=int(getattr(table, "totalsRowCount", 0) or 0),
    )
