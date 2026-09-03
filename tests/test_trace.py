"""The dependency trace: "where does this number come from?"

Only a tool that built the graph can answer that, so it gets its own tests.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from gridlint import engine, workbook
from gridlint.audit import audit

SAMPLES = Path(__file__).resolve().parent.parent / "samples"


@pytest.fixture(scope="module")
def board():
    wb = workbook.load(SAMPLES / "board-model.xlsx")
    graph = engine.build_graph(wb)
    return wb, graph, engine.recalc(wb, graph)


def test_the_trace_walks_from_a_summary_cell_back_to_its_inputs(board):
    wb, graph, computed = board
    steps = engine.trace(wb, graph, computed, ("Operating Model", 3, 26))   # cost per head
    cells = [s.cell for s in steps]
    assert cells[0] == "Operating Model!C26"
    assert "Operating Model!C25" in cells, "the full-year total should be on the path"
    assert any(s.cell == "Operating Model!C16" for s in steps), "so should a monthly total"
    assert any(s.is_input for s in steps), "the chain must reach a typed number"


def test_the_trace_labels_each_step_with_the_row_it_belongs_to(board):
    wb, graph, computed = board
    steps = engine.trace(wb, graph, computed, ("Operating Model", 3, 26))
    labels = {s.label for s in steps if s.label}
    assert "Cost per head" in labels
    assert "Full-year costs" in labels


def test_depth_increases_away_from_the_target(board):
    wb, graph, computed = board
    steps = engine.trace(wb, graph, computed, ("Operating Model", 3, 26))
    assert steps[0].depth == 0
    assert max(s.depth for s in steps) >= 2
    for a, b in zip(steps, steps[1:]):
        assert b.depth >= a.depth, "the walk must be breadth-first, so depth never goes backwards"


def test_the_trace_prefers_the_branch_that_actually_moves(board):
    wb, graph, computed = board
    changed = {("Operating Model", 3, 25), ("Operating Model", 3, 16)}
    steps = engine.trace(wb, graph, computed, ("Operating Model", 3, 26), changed=changed)
    moving = [s for s in steps if s.changed]
    assert {s.cell for s in moving} >= {"Operating Model!C25", "Operating Model!C16"}


def test_the_trace_stops_at_the_depth_it_is_given(board):
    wb, graph, computed = board
    steps = engine.trace(wb, graph, computed, ("Operating Model", 3, 26), max_depth=1)
    assert max(s.depth for s in steps) <= 1


def test_a_trace_of_an_input_cell_is_just_that_cell(board):
    wb, graph, computed = board
    steps = engine.trace(wb, graph, computed, ("Operating Model", 3, 11))
    assert len(steps) == 1 and steps[0].is_input


def test_a_cycle_does_not_make_the_trace_loop_forever(tmp_path):
    p = tmp_path / "cycle.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"
    ws["C1"] = "=A1*2"
    wb.save(p)

    m = workbook.load(p)
    g = engine.build_graph(m)
    steps = engine.trace(m, g, engine.recalc(m, g), ("Sheet", 3, 1))
    assert len(steps) <= 8
    assert len({s.cell for s in steps}) == len(steps), "each cell should appear once"


def test_the_report_attaches_a_trace_to_the_top_finding():
    report = audit(SAMPLES / "board-model.xlsx")
    top = report.findings[0]
    assert top.trace, "the finding with a measured impact should carry its trace"
    first = top.trace[0]
    assert first["cell"] == top.headline_changes[0]["cell"], \
        "the trace should start at the number a person reads, not at the defect"
    assert any(s["changed"] for s in top.trace)
    assert any(s["is_input"] for s in top.trace)


def test_a_finding_with_no_measured_impact_has_no_trace():
    report = audit(SAMPLES / "board-model.xlsx")
    unpriced = [f for f in report.findings if not f.headline_changes]
    assert unpriced, "expected at least one finding without a measured impact"
    assert all(not f.trace for f in unpriced)
