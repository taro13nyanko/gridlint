"""Engine tests: dependency graph, recalculation order, cycles, and the self-check.

`test_conformance_workbook_matches_excel` is the one that matters most: it
compares the engine against values Excel itself computed. If Gridlint's numbers
cannot be trusted, none of its findings can be either.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from gridlint import engine, workbook
from gridlint.formula.evaluator import CIRC
from gridlint.formula.values import ExcelError

SAMPLES = Path(__file__).resolve().parent.parent / "samples"


@pytest.fixture(scope="module")
def conformance():
    return workbook.load(SAMPLES / "conformance.xlsx")


@pytest.fixture(scope="module")
def board():
    return workbook.load(SAMPLES / "board-model.xlsx")


def test_conformance_workbook_matches_excel(conformance):
    """Every formula in the conformance workbook must recompute to what Excel saved."""
    sc = engine.self_check(conformance)
    assert sc.checked >= 90, "conformance fixture lost cases"
    assert not sc.unsupported, f"unsupported formulas: {sc.unsupported}"
    assert sc.mismatches == [], f"engine disagrees with Excel: {sc.mismatches}"
    assert sc.agreement == 1.0


def test_board_model_matches_excel(board):
    sc = engine.self_check(board)
    assert sc.mismatches == []
    assert sc.trustworthy


def test_clean_model_matches_excel():
    sc = engine.self_check(workbook.load(SAMPLES / "clean-model.xlsx"))
    assert sc.mismatches == []


def test_graph_orders_dependencies_before_dependents(board):
    g = engine.build_graph(board)
    position = {k: i for i, k in enumerate(g.order)}
    for key, prec in g.precedents.items():
        for p in prec.cells:
            if p in position and key in position and p != key:
                assert position[p] < position[key] or p in {c for cy in g.cycles for c in cy}, \
                    f"{p} evaluated after its dependent {key}"


def test_downstream_finds_transitively_affected_cells(board):
    g = engine.build_graph(board)
    seed = ("Operating Model", 3, 11)          # first cost line, January
    down = g.downstream([seed])
    assert ("Operating Model", 3, 16) in down          # the monthly total
    assert len(down) > 5


def test_overrides_change_dependent_values(board):
    g = engine.build_graph(board)
    base = engine.recalc(board, g)
    key_total = ("Operating Model", 3, 16)
    bumped = engine.recalc(board, g, overrides={("Operating Model", 3, 11): 1_000_000.0})
    assert bumped[key_total] != base[key_total]


def test_formula_override_does_not_mutate_the_original_workbook(board):
    g = engine.build_graph(board)
    before = board.get("Operating Model", 3, 16).formula
    engine.recalc(board, g, formula_overrides={("Operating Model", 3, 16): "=SUM(C11:C15)"})
    assert board.get("Operating Model", 3, 16).formula == before


def test_circular_reference_is_detected_and_reported(tmp_path):
    import openpyxl

    p = tmp_path / "cycle.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"
    ws["C1"] = "=A1*2"
    wb.save(p)

    m = workbook.load(p)
    g = engine.build_graph(m)
    assert g.cycles, "cycle not detected"
    members = {c for cycle in g.cycles for c in cycle}
    assert ("Sheet", 1, 1) in members and ("Sheet", 2, 1) in members
    computed = engine.recalc(m, g)
    assert isinstance(computed[("Sheet", 1, 1)], ExcelError)
    assert computed[("Sheet", 1, 1)].code == CIRC


def test_long_reference_chain_does_not_hit_recursion_limit(tmp_path):
    import openpyxl

    p = tmp_path / "chain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1
    for r in range(2, 3002):
        ws.cell(row=r, column=1).value = f"=A{r - 1}+1"
    wb.save(p)

    m = workbook.load(p)
    computed = engine.recalc(m)
    assert computed[("Sheet", 1, 3001)] == 3001.0


def test_workbook_without_cached_values_is_reported_not_crashed(tmp_path):
    import openpyxl

    p = tmp_path / "nocache.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "=1+1"
    wb.save(p)

    m = workbook.load(p)
    assert not m.has_cached_values
    sc = engine.self_check(m)
    assert sc.checked == 0 and not sc.trustworthy
    assert "no cached values" in sc.summary()


def test_values_equal_tolerates_float_noise_but_not_real_differences():
    assert engine.values_equal(1_000_000.0, 1_000_000.0000001)
    assert not engine.values_equal(1_000_000.0, 1_000_001.0)
    assert engine.values_equal(ExcelError("#DIV/0!"), ExcelError("#DIV/0!"))
    assert not engine.values_equal(ExcelError("#DIV/0!"), ExcelError("#VALUE!"))
    assert not engine.values_equal(ExcelError("#DIV/0!"), 0.0)
