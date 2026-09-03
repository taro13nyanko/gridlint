"""Command line behaviour, including the exit codes CI depends on."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from gridlint.cli import main

SAMPLES = Path(__file__).resolve().parent.parent / "samples"


def run(argv, capsys) -> tuple[int, str]:
    code = main(argv)
    return code, capsys.readouterr().out


def test_check_exits_one_when_a_critical_defect_is_found(capsys):
    code, out = run(["check", str(SAMPLES / "board-model.xlsx")], capsys)
    assert code == 1
    assert "Total leaves out a row of data" in out
    assert "verified by recalculation" in out
    assert "Runway (months): 38.6 becomes 5.2" in out


def test_check_exits_zero_on_a_clean_workbook(capsys):
    code, out = run(["check", str(SAMPLES / "clean-model.xlsx")], capsys)
    assert code == 0
    assert "No defects found" in out


def test_fail_on_warning_catches_the_softer_findings(capsys):
    clean = str(SAMPLES / "clean-model.xlsx")
    assert run(["check", clean, "--fail-on", "warning"], capsys)[0] == 0
    board = str(SAMPLES / "board-model.xlsx")
    assert run(["check", board, "--fail-on", "warning"], capsys)[0] == 1


def test_json_output_is_parseable_and_complete(capsys):
    code, out = run(["check", str(SAMPLES / "board-model.xlsx"), "--json"], capsys)
    assert code == 1
    d = json.loads(out)
    assert d["engine"]["trustworthy"]
    assert d["findings"][0]["fix"]["new_formula"] == "=SUM(C11:C15)"
    assert d["findings"][0]["headline_changes"]


def test_explain_uses_the_same_context_as_the_web_app(capsys):
    """A fixture recorded once must replay from either entry point."""
    code, out = run(["check", str(SAMPLES / "board-model.xlsx"), "--explain"], capsys)
    assert code == 1
    assert "note:" in out, "recorded explanations did not replay through the CLI"
    assert "Review note" in out


def test_verify_reports_engine_agreement(capsys):
    code, out = run(["verify", str(SAMPLES / "conformance.xlsx")], capsys)
    assert code == 0
    assert "93/93" in out or "reproduced" in out


def test_rules_lists_every_code(capsys):
    code, out = run(["rules"], capsys)
    assert code == 0
    for rule in ("R001", "R002", "R005", "R012", "R013"):
        assert rule in out


def test_fix_dry_run_shows_edits_without_writing(tmp_path, capsys):
    src = SAMPLES / "runway.xlsx"
    code, out = run(["fix", str(src), "--dry-run"], capsys)
    assert code == 0
    assert "=SUM(C11:C14)  ->  =SUM(C11:C15)" in out
    assert not list(SAMPLES.glob("*-fixed.xlsx")), "dry run must not write a file"


def test_fix_writes_a_corrected_copy_and_leaves_the_original_alone(tmp_path, capsys):
    src = SAMPLES / "runway.xlsx"
    before = src.read_bytes()
    out_path = tmp_path / "fixed.xlsx"
    code, _ = run(["fix", str(src), "--out", str(out_path)], capsys)
    assert code == 0
    assert src.read_bytes() == before, "the original workbook must not be touched"

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Model"]
    assert ws["C16"].value == "=SUM(C11:C15)"
    assert ws["H16"].value == "=SUM(H11:H15)"
    assert "was: =SUM(C11:C14)" in ws["C16"].comment.text
    wb.close()


def test_only_limits_the_fix_to_one_rule(tmp_path, capsys):
    out_path = tmp_path / "r002.xlsx"
    code, out = run(["fix", str(SAMPLES / "board-model.xlsx"), "--only", "R002",
                     "--out", str(out_path)], capsys)
    assert code == 0
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Operating Model"]
    assert ws["K13"].value == "=J13*(1+$B$13)"
    assert ws["C16"].value == "=SUM(C11:C14)", "R001 was not selected and must be left alone"
    wb.close()


def test_a_missing_file_is_reported_not_traced(capsys):
    assert main(["check", "no-such-file.xlsx"]) == 2


def test_version_flag(capsys):
    with pytest.raises(SystemExit) as e:
        main(["--version"])
    assert e.value.code == 0
