"""The things a real workbook contains that a toy one does not.

Every case here comes from a formula that broke Gridlint the first time it met
one: newer functions stored with an `_xlfn.` prefix, array idioms, dates as
serial numbers, structured table references, links into other workbooks, and
rows somebody hid.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.table import Table, TableStyleInfo

from gridlint import engine, workbook
from gridlint.audit import audit
from gridlint.formula.evaluator import evaluate
from gridlint.formula.parser import parse_formula, parse_table_ref
from gridlint.formula.tokenizer import tokenize
from gridlint.formula.values import ExcelError, from_serial, to_serial

SAMPLES = Path(__file__).resolve().parent.parent / "samples"


def rules_found(report) -> set[str]:
    return {f.rule for f in report.findings}


def find(report, rule: str):
    return next((f for f in report.findings if f.rule == rule), None)


# ------------------------------------------------------- functions Excel added later
def test_newer_functions_are_stored_with_a_marker_prefix():
    """Excel writes MAXIFS into the file as _xlfn.MAXIFS. Reading the file means stripping it."""
    toks = tokenize("=_xlfn.MAXIFS(A1:A9,B1:B9,1)")
    assert toks[0].value == "MAXIFS"
    node = parse_formula('=_xlfn.IFS(A1>1,"a",TRUE,"b")')
    assert node.name == "IFS"


def test_the_function_set_covers_what_operating_models_are_built_from():
    from gridlint.formula.functions import FUNCTIONS

    for name in ("SUMIFS", "COUNTIFS", "AVERAGEIFS", "MAXIFS", "MINIFS", "SUMPRODUCT",
                 "XLOOKUP", "IFS", "SWITCH", "CHOOSE", "HLOOKUP",
                 "TODAY", "DATE", "YEAR", "MONTH", "EOMONTH", "EDATE", "DATEDIF",
                 "NPV", "IRR", "PMT", "PV", "FV",
                 "MOD", "CEILING", "FLOOR", "MEDIAN", "STDEV", "LARGE", "SMALL",
                 "TEXTJOIN", "SUBSTITUTE", "FIND", "SEARCH"):
        assert name in FUNCTIONS, f"{name} is missing"


# ------------------------------------------------------------------------ dates
def test_dates_convert_to_the_serial_number_excel_stores():
    assert to_serial(dt.date(2026, 9, 3)) == 46268
    assert to_serial(dt.date(2000, 1, 1)) == 36526
    assert from_serial(46268).date() == dt.date(2026, 9, 3)


def test_the_1900_leap_year_bug_is_preserved():
    """Excel believes 1900 was a leap year. Anything reading its files has to agree."""
    assert to_serial(dt.date(1900, 1, 1)) == 1
    assert to_serial(dt.date(1900, 3, 1)) == 61        # serial 60 is the day that never was


def test_a_date_cell_and_a_serial_number_compare_equal(tmp_path):
    """The file gives back a datetime; the engine computes the number under it."""
    assert engine.values_equal(dt.datetime(2026, 9, 3), 46268.0)
    assert engine.values_equal(dt.date(2026, 9, 3), 46268)
    assert not engine.values_equal(dt.date(2026, 9, 3), 46269)


# ----------------------------------------------------------------- array idioms
class _Grid:
    def __init__(self, cells):
        self.cells = cells

    def cell(self, sheet, col, row):
        from gridlint.formula.values import BLANK
        return self.cells.get((col, row), BLANK)

    def used_bounds(self, sheet):
        return max(c for c, _r in self.cells), max(r for _c, r in self.cells)

    def sheet_exists(self, sheet):
        return True

    def defined_name(self, name):
        return None

    def table_range(self, table, column):
        return None


def test_sumproduct_with_a_comparison_array():
    """SUMPRODUCT((range="x")*range) predates SUMIFS and is everywhere in real models."""
    cells = {}
    for i, (region, amount) in enumerate([("North", 10.0), ("South", 20.0), ("North", 30.0)], start=1):
        cells[(1, i)] = region
        cells[(2, i)] = amount
    got = evaluate(parse_formula('=SUMPRODUCT((A1:A3="North")*B1:B3)'), "S", _Grid(cells))
    assert got == 40.0


def test_comparison_against_a_range_broadcasts():
    cells = {(1, 1): 5.0, (1, 2): 15.0, (1, 3): 25.0}
    got = evaluate(parse_formula("=SUMPRODUCT((A1:A3>10)*1)"), "S", _Grid(cells))
    assert got == 2.0


# ------------------------------------------------------------ table references
def test_structured_reference_parses_into_table_and_column():
    t = parse_table_ref("Sales[[#Data],[Amount]]")
    assert (t.table, t.column) == ("Sales", "Amount")
    assert parse_table_ref("Table1[Amount]").column == "Amount"


def _workbook_with_table(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    for region, amount in (("North", 100), ("South", 250), ("North", 400)):
        ws.append([region, amount])
    table = Table(displayName="Sales", ref="A1:B4")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
    ws.add_table(table)
    ws["D1"] = "=SUM(Sales[Amount])"
    ws["D2"] = '=SUMIFS(Sales[Amount],Sales[Region],"North")'
    wb.save(path)
    return path


def test_a_table_reference_resolves_to_the_column_it_names(tmp_path):
    m = workbook.load(_workbook_with_table(tmp_path / "table.xlsx"))
    assert "sales" in m.tables
    computed = engine.recalc(m)
    assert computed[("Sheet", 4, 1)] == 750.0          # SUM(Sales[Amount])
    assert computed[("Sheet", 4, 2)] == 500.0          # only the North rows


def test_a_table_reference_feeds_the_dependency_graph(tmp_path):
    m = workbook.load(_workbook_with_table(tmp_path / "table2.xlsx"))
    g = engine.build_graph(m)
    assert ("Sheet", 4, 1) in g.downstream([("Sheet", 2, 2)]), \
        "editing a table cell must be seen to change the total"


# --------------------------------------------------------------- external links
def test_an_external_link_is_reported_rather_than_guessed(tmp_path):
    p = tmp_path / "ext.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=[1]Budget!B4*2"
    ws["A2"] = "=[1]Budget!B5"
    ws["A3"] = "=A1+A2"
    wb.save(p)

    report = audit(p)
    f = find(report, "R014")
    assert f is not None
    assert f.evidence["count"] == 2
    assert f.group_size == 2
    assert "R005" not in rules_found(report), \
        "an external link must not also be reported as an unrecognised function"


def test_a_quoted_external_path_is_treated_as_a_sheet_reference():
    node = parse_formula("='[Budget 2026.xlsx]Sheet1'!A1")
    assert getattr(node, "sheet", "").startswith("[Budget")


# ---------------------------------------------------------------- hidden rows
def test_a_hidden_row_inside_a_total_is_reported_with_its_contribution(tmp_path):
    p = tmp_path / "hidden.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate([100, 200, 300, 400], start=1):
        ws.cell(row=i, column=1, value=v)
    ws["A5"] = "=SUM(A1:A4)"
    ws.row_dimensions[3].hidden = True
    wb.save(p)

    f = find(audit(p), "R015")
    assert f is not None
    assert f.evidence["hidden_rows"] == [3]
    assert f.evidence["hidden_contribution"] == 300.0


def test_no_hidden_row_finding_when_nothing_is_hidden(tmp_path):
    p = tmp_path / "visible.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate([100, 200, 300], start=1):
        ws.cell(row=i, column=1, value=v)
    ws["A4"] = "=SUM(A1:A3)"
    wb.save(p)
    assert "R015" not in rules_found(audit(p))


# ------------------------------------------------------------------- odds and ends
def test_implicit_intersection_marker_is_ignored():
    """Newer Excel writes =@A1 for what used to be =A1."""
    node = parse_formula("=@A1")
    assert getattr(node, "col", None) == 1


def test_the_conformance_workbook_still_matches_excel_after_all_of_this():
    sc = engine.self_check(workbook.load(SAMPLES / "conformance.xlsx"))
    assert sc.checked >= 150
    assert sc.mismatches == []
    assert not sc.unsupported
